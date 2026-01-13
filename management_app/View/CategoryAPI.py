from rest_framework.response import Response
from rest_framework import status
from management_app.serializer.CategorySerializer import *
from ..models import *
# from sales_client_app.paginations import WebProductPaginationClass
from django.utils.text import slugify
from django.conf import settings
from rest_framework.views import APIView
from django.db.models import Q
import csv
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from ..pagination import ListPagination
from rest_framework.permissions import IsAuthenticated, IsAdminUser


class CategoryAPI(APIView):
    permission_classes = [IsAuthenticated]

    def get(self,request):
        user = request.user

        categories = CategoryModel.objects.filter(is_active=True,admin_user=user).order_by('-id')
        search = request.query_params.get('search','')
        category_list=[]
        paginator = ListPagination()
        
        if search:
            categories = categories.filter(Q(name__icontains=search)|
                                           Q(is_active__icontains = search))
            
        for category in categories:
            if category.is_root():
                serializer = CategorySerializer(category).data
                category_list.append(serializer)
            else:
                continue
         
      
        paginated_categories = paginator.paginate_queryset(category_list,request)

       
        all_records = request.query_params.get('all', False)
        if all_records and all_records in ['1', 'true', 'True', True]:
            paginated_categories = None

        return paginator.get_paginated_response(paginated_categories)
    
    def post(self,request):
        data = request.data
        
        serializer = CategorySerializer(data=data,context={'request': request})
        if serializer.is_valid():
            serializer.save()
            
            return Response({'status':True,'data':serializer.data,'message':'Category successfully added'})
        else:
             return Response({'status':False,'errors':serializer.errors})
        
    def patch(self,request,id):
    

        try:
            category = CategoryModel.objects.get(id=id,admin_user=request.user)
            serializer = CategorySerializer(category,data=request.data,partial=True,context={'request': request})
            if serializer.is_valid():
                serializer.save()
                return Response({'status':True,'data':serializer.data,'message':'Category successfully updated'})
            else:
                return Response({'status':False,'errors':serializer.errors})
        except CategoryModel.DoesNotExist:
            return Response({'status':False,'message':'Category not available'},status=status.HTTP_400_BAD_REQUEST)
        
    def delete(self,request,id):
        user = request.user

        try :
            category = CategoryModel.objects.get(id=id,admin_user=user)
            category.delete()
            return Response({'status':True,'message':'Category successfully deleted'})
        except CategoryModel.DoesNotExist:
           return Response({'status':False,'message':'Category not available'},status=status.HTTP_400_BAD_REQUEST)
                 

            
class SubCategoryAPI(APIView):
    def get(self,request,id=None):
        user = request.user
        sub_categories = CategoryModel.objects.filter(is_active=True,admin_user=user).order_by('-id')
        search = request.query_params.get("search","")
        main_category = request.query_params.get('category','')
        all_flag = request.query_params.get("all", "false").lower() == "true" 
        sub_categories_list=[]
        if id :
            parent_category= CategoryModel.objects.get(id=id,admin_user=user)
            sub_categories = parent_category.get_children()
            for sub_category in sub_categories:
                serializer  = SubCategorySerializer(sub_category).data
                sub_categories_list.append(serializer)
            return Response({'status':True,'sub_categories':sub_categories_list,'message':'Categories successfully retrieved'})
        
        if search:
            sub_categories = sub_categories.filter(Q(name__icontains = search) )

        if main_category:
            sub_categories = sub_categories.filter(name=main_category).first().get_children()
            

        for sub_category in sub_categories:
            if sub_category.get_parent():
                serialized  = SubCategorySerializer(sub_category).data
                sub_categories_list.append(serialized)
            else:
                continue
        
        if all_flag:
            return Response({
                'status': True,
                'sub_categories': sub_categories_list,
                'message': 'All subcategories retrieved (no pagination)'
            })

        paginator = ListPagination()        
        paginated_sub_categories = paginator.paginate_queryset(sub_categories_list,request)

        return paginator.get_paginated_response(paginated_sub_categories)
    
    def post(self,request):
        data = request.data
        
        serializer = SubCategorySerializer(data=data,context={'request': request})

        if serializer.is_valid():
            serializer.save()
            return Response({'status':True,'data':serializer.data,'message':'Sub Category successfully added'})
        else:
             return Response({'status':False,'errors':serializer.errors})
        
    def patch(self,request,id):
        try:
              
           sub_category = CategoryModel.objects.get(id=id,admin_user=request.user)
           serializer = SubCategorySerializer(sub_category,data=request.data,partial=True,context={'request': request})
           if serializer.is_valid():
               serializer.save()
               return Response({'status':True,'data':serializer.data,'message':'Sub Category successfully updated'})
           else:
                return Response({'status':False,'errors':serializer.errors})
           
             
        except CategoryModel.DoesNotExist:
            return Response({'status':False,'message':'Sub Category not available'},status=status.HTTP_400_BAD_REQUEST)
        
    def delete(self,request,id):
        try :
            sub_category = CategoryModel.objects.get(id=id,admin_user=request.user)
            sub_category.delete()
            return Response({'status':True,'message':'Category successfully deleted'})
        except CategoryModel.DoesNotExist:
           return Response({'status':False,'message':'Category not available'},status=status.HTTP_400_BAD_REQUEST)


class Export_categories_excel(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, *args, **kwargs):
        user = request.user
        # if not request.user.is_authenticated:
        #     return Response(
        #         {"status": False, "message": "Authentication credentials were not provided."},
        #         status=status.HTTP_401_UNAUTHORIZED,
        #     )


        export_dir = os.path.join(settings.MEDIA_ROOT, "export", "categories")
        os.makedirs(export_dir, exist_ok=True)


        file_name = f'categories{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        file_path = os.path.join(export_dir, file_name)


        headers = ["ID", "Name", "Is Active"]

        qs = CategoryModel.objects.filter(admin_user=user)


        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Categories"

        # Write header (bold)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Write rows
        for c in qs:
            if c.is_root():  # keep your condition
                ws.append([
                    c.id,
                    c.name,
                    "Yes" if c.is_active else "No",
                ])


        wb.save(file_path)


        file_uri = os.path.join(settings.MEDIA_URL, "export", "categories", file_name)
        absolute_file_uri = request.build_absolute_uri(file_uri)

        return Response({
            "status": True,
            "file_uri": absolute_file_uri,
            "message": "Categories successfully exported"
        }, status=200)

class Export_sub_categories_excel(APIView):
    permission_classes = [IsAuthenticated]
    def get(self,request):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sub Categories"
        
        user = request.user

        export_dir = os.path.join(settings.MEDIA_ROOT, "export", "sub categories")
        os.makedirs(export_dir, exist_ok=True)

        file_name = f'sub_categories{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        file_path = os.path.join(export_dir, file_name)

        # header row
        sheet.append(["ID", "Name",'Main Cateogory','Is Active'])
        for cell in sheet[1]:  # first row
            cell.font = Font(bold=True)

        qs = CategoryModel.objects.filter(admin_user=user)

        # data rows
        for category in qs:
            if category.get_parent():
                sheet.append([
                category.id,
                category.name,
                category.get_parent().name,
                "Yes" if category.is_active else "No",
                ])
            else:
                continue

        workbook.save(file_path)


        file_uri = os.path.join(settings.MEDIA_URL, "export", "sub categories", file_name)
        absolute_file_uri = request.build_absolute_uri(file_uri)

        return Response({
            "status": True,
            "file_uri": absolute_file_uri,
            "message": "Sub Categories successfully exported"
        }, status=200)
    
class HomeCategoryView(APIView):
    def get(self,request):
        user = request.user
        home_categories = HomeCategoryModel.objects.filter(admin_user=user)
        data = []
        for cat in home_categories:
            data.append({
               "id" :cat.id,
               "name" : cat.name
            }
            )
        return Response(data)