from rest_framework.response import Response
from rest_framework import status
from management_app.serializer.OfferSerializer import *
from ..models import *
# from sales_client_app.paginations import WebProductPaginationClass
from django.utils.text import slugify
from rest_framework.views import APIView
from django.db.models import Q
from decimal import Decimal, InvalidOperation
import csv
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from django.utils.timezone import make_naive
from openpyxl.drawing.image import Image as XLImage
import os
from ..pagination import ListPagination
from django.conf import settings




class  OfferView(APIView):
    
    def get(self, request):
        offers = OfferModel.objects.all()
        search = request.query_params.get('search','')
        if search:
            offers = OfferModel.objects.filter(Q(title__icontains=search) |
                                               Q(roles__type__icontains=search) |
                                               Q(description__icontains=search)
                                               )
        paginator = ListPagination()
        paginated_offers = paginator.paginate_queryset(offers,request)
        serializer = OfferSerializer(paginated_offers, many=True)
        return paginator.get_paginated_response(serializer.data)

    def post(self, request):
        serializer = OfferSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'status':True,'data':serializer.data}, status=status.HTTP_201_CREATED)
        return Response({'status':False,'errors':serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
    
    def patch(self,request,id):
        try:
            online_offer = OfferModel.objects.get(id=id)
            serializer = OfferSerializer(online_offer,data=request.data,partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({'status':True,'data':serializer.data,'message':'Offer Successfully updated'})
            return Response({'status':False,'errors':serializer.errors})
        except OfferModel.DoesNotExist:
            return Response({'status':False,'message':'Offer not available'},status=status.HTTP_400_BAD_REQUEST)
    

    def delete(self,request,id):
        try:
            online_offer = OfferModel.objects.get(id=id)
            online_offer.delete()
            return Response({'status':True,'message':'Offer Successfully deleted'})

        except OfferModel.DoesNotExist:
           return Response({'status':False,'message':'Offer not available'},status=status.HTTP_400_BAD_REQUEST)


class Export_offers_excel(APIView):
    
    def get(self,request):
        if not request.user.is_authenticated:
            return Response(
                {"status": False, "message": "Authentication credentials were not provided."},
                status=status.HTTP_401_UNAUTHORIZED,
            )

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Offers"

        export_dir = os.path.join(settings.MEDIA_ROOT,"export","offers")
        os.makedirs(export_dir,exist_ok=True)

        file_name = 'offers.xlsx'
        file_path = os.path.join(export_dir,file_name)

        # header row
        sheet.append(["ID", "Title",'Role','Description'])
        for cell in sheet[1]:  # first row
            cell.font = Font(bold=True)

        # data rows
        for online_offer in OfferModel.objects.all():
            sheet.append([online_offer.id, online_offer.title,online_offer.roles.name if online_offer.roles else '',online_offer.description or ''])

        workbook.save(file_path)

        file_uri = os.path.join(settings.MEDIA_URL,'export','offers',file_name)
        absolute_file_uri = request.build_absolute_uri(file_uri)


        return Response({
            "status": True,
            "file_uri": absolute_file_uri,
            "message": "Offers successfully exported"
        }, status=200)
    

class OnlinePaymentOfferView(APIView):

    def get(self, request):
        online_offers = OnlinePaymentOfferModel.objects.all()
        search = request.query_params.get('search','')
        if search:
            online_offers = OnlinePaymentOfferModel.objects.filter(Q(start_price__icontains=search) |
                                               Q(end_price__icontains=search) |
                                               Q(percentage_off__icontains=search)
                                               )
        serializer = OnlinePaymentOfferSerializer(online_offers, many=True)
        return Response({'status':True,'data':serializer.data},status=status.HTTP_200_OK)

    def post(self, request):
        serializer = OnlinePaymentOfferSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'status':True,'data':serializer.data}, status=status.HTTP_201_CREATED)
        return Response({'status':False,'errors':serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
    
    def patch(self,request,id):
        try:
            online_offer = OnlinePaymentOfferModel.objects.get(id=id)
            serializer = OnlinePaymentOfferSerializer(online_offer,data=request.data,partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({'status':True,'data':serializer.data,'message':'Online Offer Successfully updated'})
            return Response({'status':False,'errors':serializer.errors})
        except OfferModel.DoesNotExist:
            return Response({'status':False,'message':'Online Offer not available'},status=status.HTTP_400_BAD_REQUEST)
    

    def delete(self,request,id):
        try:
            online_offer = OnlinePaymentOfferModel.objects.get(id=id)
            online_offer.delete()
            return Response({'status':True,'message':'Online Offer Successfully deleted'})

        except OfferModel.DoesNotExist:
           return Response({'status':False,'message':'Online Offer not available'},status=status.HTTP_400_BAD_REQUEST)
        


class Export_online_offers_excel(APIView):
    
    def get(self,request):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Online Offers"

        export_dir = os.path.join(settings.MEDIA_ROOT,"export","online offers")
        os.makedirs(export_dir,exist_ok=True)

        file_name = 'online_offers.xlsx'
        file_path = os.path.join(export_dir,file_name)
        # header row
        sheet.append(["ID", "Start Price",'End Price','Percentage Off'])
        for cell in sheet[1]:  # first row
            cell.font = Font(bold=True)

        # data rows
        for online_offer in OnlinePaymentOfferModel.objects.all():
            sheet.append([online_offer.id, online_offer.start_price,online_offer.end_price,online_offer.percentage_off or ''])

        workbook.save(file_path)
        file_uri = os.path.join(settings.MEDIA_URL,'export','online offers',file_name)
        absolute_file_uri = request.build_absolute_uri(file_uri)


        return Response({
            "status": True,
            "file_uri": absolute_file_uri,
            "message": "Online Offers successfully exported"
        }, status=200)

    
class OfferSliderView(APIView):

    def get(self, request):
        offer_slider = OfferSliderModel.objects.all()
        search = request.query_params.get('search','')
        if search:
            offer_slider = OfferSliderModel.objects.filter(Q(banner_number__icontains=search)
                                               )
        serializer = OfferSliderSerializer(offer_slider, many=True)
        return Response({'status':True,'data':serializer.data},status=status.HTTP_200_OK)

    def post(self, request):
        serializer = OfferSliderSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'status':True,'data':serializer.data,'message':'Offer Slider Image Data Successfully created'}, status=status.HTTP_201_CREATED)
        return Response({'status':False,'errors':serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
    
    def patch(self,request,id):
        try:
            offer_slider = OfferSliderModel.objects.get(id=id)
            serializer = OfferSliderSerializer(offer_slider,data=request.data,partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({'status':True,'data':serializer.data,'message':'Offer Slider Image Data Successfully updated'},status=status.HTTP_200_OK)
            return Response({'status':False,'errors':serializer.errors})
        except OfferModel.DoesNotExist:
            return Response({'status':False,'message':'Offer Slider Image not available'},status=status.HTTP_400_BAD_REQUEST)
    

    def delete(self,request,id):
        try:
            online_offer = OfferSliderModel.objects.get(id=id)
            online_offer.delete()
            return Response({'status':True,'message':'Offer Slider Image Data Successfully deleted'})

        except OfferModel.DoesNotExist:
           return Response({'status':False,'message':'Offer Slider Image Data not available'},status=status.HTTP_400_BAD_REQUEST)
        

class Export_offer_slider_image_excel(APIView):
    
    def get(self,request):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Offer Slider Image"
        row_num = 2

        export_dir = os.path.join(settings.MEDIA_ROOT,"export","offer slider")
        os.makedirs(export_dir,exist_ok=True)

        file_name = 'offer_slider.xlsx'
        file_path = os.path.join(export_dir,file_name)
        # header row
        sheet.append(["ID", "Image",'Banner No.','Created At'])
        for cell in sheet[1]:  # first row
            cell.font = Font(bold=True)

        # data rows
        for offer_slider in OfferSliderModel.objects.all().order_by('id'):
            sheet.append([offer_slider.id,'',offer_slider.banner_number or '', make_naive(offer_slider.created_at).strftime("%Y-%m-%d, %H:%M:%S")])
            if offer_slider.image:
                image_path = offer_slider.image.path  # local filesystem path
                if os.path.exists(image_path):
                    img = XLImage(image_path)
                    img.width = 40   # resize
                    img.height = 18
                    sheet.add_image(img, f"B{row_num}")  # put in column B
            row_num += 1

        workbook.save(file_path)
        file_uri = os.path.join(settings.MEDIA_URL,'export','offer slider',file_name)
        absolute_file_uri = request.build_absolute_uri(file_uri)


        return Response({
            "status": True,
            "file_uri": absolute_file_uri,
            "message": "Offer slider image data successfully exported"
        }, status=200)

