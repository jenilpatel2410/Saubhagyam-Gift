from rest_framework.response import Response
from rest_framework import status
from management_app.serializer.ProductSerializer import *
from ..models import *
# from sales_client_app.paginations import WebProductPaginationClass
from django.utils.text import slugify
from rest_framework.views import APIView
from django.db.models import Q
from decimal import Decimal, InvalidOperation
import csv
import openpyxl, io
from django.http import HttpResponse
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage
from ..pagination import ListPagination
from io import BytesIO
from datetime import datetime
from django.conf import settings
from django.template.loader import get_template
from xhtml2pdf import pisa
import ast  
from urllib.parse import urljoin


class ProductAPI(APIView):
    def get(self,request,id=None):
        search = request.query_params.get('search','').strip()
        cat_id = request.query_params.get('cat_id','')
        company_id = request.query_params.get('company_id','')
        home_category = request.query_params.get('home_category','')
        all_flag = request.query_params.get("all", "false").lower() == "true" 
        products = ProductModel.objects.filter(is_active=True).order_by('-id')

        if id:
            product = products.get(id=id)
            serializer = ProductDetailSerializer(product)
            return Response({'status':True,'data':serializer.data,'message':'Product details retrieved successfully'})
        
        if home_category:
            products = products.filter(home_categories__name=home_category)
        
        if company_id:
            products_count = products.filter(category__id=company_id).count()
            products = products.filter(company__id=company_id)
        
        if cat_id:
            products = products.filter(category__id = cat_id)
            products_count1 = products.filter(sub_category__id = cat_id).count()
        
        if search:
            search_params = [s.strip() for s in search.split(',') if s.strip()]
            q = Q()

            for term in search_params:
                sub_q = (
                    Q(name__icontains=term) |
                    Q(item_code__icontains=term) |
                    Q(category__name__icontains=term) |
                    Q(company__code__icontains=term) |
                    Q(company__name__icontains=term) |
                    Q(unit__icontains=term) |
                    Q(short_name__icontains=term) |
                    Q(model__icontains=term) |
                    Q(warehouse_section__icontains=term) |
                    Q(description__icontains=term) |
                    Q(limited_stock__icontains=term) |
                    Q(out_of_stock__icontains=term) |
                    Q(brand__name__icontains=term)
                )

                # Try converting to decimal
                try:
                    search_decimal = Decimal(term)
                    sub_q |= (
                        Q(product_price=search_decimal) |
                        Q(cost=search_decimal) |
                        Q(weight=search_decimal)
                    )
                except (InvalidOperation, ValueError):
                    pass

                # Add partial result to main Q
                q |= sub_q   # <= If you want ANY term match (change based on your requirement)

            products = products.filter(q)
        
        if all_flag:
            serializer = ProductListSerializer(products,many=True)
            return Response({
                'status': True,
                'Products': serializer.data,
                'message': 'All subcategories retrieved (no pagination)'
            })
        
        paginator = ListPagination()
        paginated_products = paginator.paginate_queryset(products,request)

        serializer = ProductListSerializer(paginated_products,many=True)
        return paginator.get_paginated_response(serializer.data)
    
    def post(self,request):
        data = request.data
        serializer = ProductSerializer(data=request.data, context={"request": request})
        if serializer.is_valid():
            serializer.save()
            return Response({'status':True,'data':serializer.data,'message':'Product successfully added'},status=status.HTTP_200_OK)
        return Response({'status':False,'errors':serializer.errors},status=status.HTTP_400_BAD_REQUEST)
    
    def patch(self,request,id):
        try:
            product = ProductModel.objects.get(id=id)
            data = request.data.copy()
            if 'company' not in data:
                data['company'] = None
            serializer = ProductSerializer(product,data=data,partial=True, context={"request": request})
            if serializer.is_valid():
                serializer.save()
                return Response({'status':True,'data':serializer.data,'message':'Product Successfully updated'},status=status.HTTP_200_OK)
            return Response({'status':False,'errors':serializer.errors},status=status.HTTP_400_BAD_REQUEST)
        
        except ProductModel.DoesNotExist:
            return Response({'status':False,'message':'Product not Available'},status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self,request,id):
        try:
            product = ProductModel.objects.get(id=id)
            product.delete()
            return Response({'status':True,'message':'Product Deleted Successfully'},status=status.HTTP_200_OK)

        
        except ProductModel.DoesNotExist:
            return Response({'status':False,'message':'Product not Available'},status=status.HTTP_400_BAD_REQUEST)
        


class ProductExcelImportAPI(APIView):
    def post(self, request):
        serializer = ProductImportSerializer(data=request.data)
        if serializer.is_valid():
            file = serializer.validated_data['file']
            filename = file.name.lower()

            created = 0
            updated = 0
            try:
                rows = []
                headers = []

                # --- Excel ---
                if filename.endswith(".xlsx"):
                    workbook = openpyxl.load_workbook(file)
                    sheet = workbook.active

                    headers = [str(cell).strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        row_dict = dict(zip(headers, row))
                        rows.append(row_dict)

                # --- CSV ---
                elif filename.endswith(".csv"):
                    decoded_file = file.read().decode("utf-8")
                    io_string = io.StringIO(decoded_file)
                    reader = csv.DictReader(io_string)  # DictReader automatically uses headers
                    for row in reader:
                        row_dict = {k.strip(): v for k, v in row.items()}
                        rows.append(row_dict)

                else:
                    return Response(
                        {"status": False, "error": "Unsupported file format. Upload .xlsx or .csv"},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # --- Process rows ---
                for row in rows:
                    name = (row.get("name") or "").strip()
                    if not name:
                        continue  # skip rows without name

                    defaults = {
                        "name": name,
                        "short_name": row.get("short_name", "") or "",
                        "warehouse_section": row.get("warehouse_section", "") or "",
                        "description": row.get("description", "") or "",
                        "product_price": row.get("product_price") or 0,
                        "cost": row.get("cost") or 0,
                        "weight": row.get("weight") or 0,
                        "item_code": row.get("item_code", "") or "",
                        "retailer_price": row.get("retailer_price") or 0,
                        "distributer_price": row.get("wholesaler_price") or 0,
                        "gst": row.get("gst") or 0,
                        "warranty": row.get("warranty", "") or "",
                        "unit": row.get("unit", "") or "",
                        "limited_stock": "Yes" if str(row.get("limited_stock")).lower() in ["yes", "true", "1"] else "No",
                        "out_of_stock": "Yes" if str(row.get("out_of_stock")).lower() in ["yes", "true", "1"] else "No",
                    }

                    # --- Brand ---
                    brand_name = row.get("brand_name")
                    if brand_name:
                        brand, _ = BrandModel.objects.get_or_create(name=brand_name.strip())
                        defaults["brand"] = brand
                    else:
                        defaults["brand"] = None

                    company_name = row.get("company_name")
                    if company_name:
                        normalized_name = company_name.strip()
                        company = CompanyModel.objects.filter(name__iexact=normalized_name).first()
                        if not company:
                            company = CompanyModel.objects.create(name=normalized_name)
                        defaults["company"] = company
                    else:
                        defaults["company"] = None
                    # --- Create/Update ---
                    product = ProductModel.objects.filter(name__iexact=name).first()
                    if product:
                        for field, value in defaults.items():
                            setattr(product, field, value)
                        product.save()
                        created_flag = False
                    else:
                        product = ProductModel.objects.create(**defaults)
                        created_flag = True

                    # --- Categories ---
                    category_names = row.get("category") or ""
                    sub_category_names = row.get("sub_category") or ""

                    category_list = []
                    for cat in str(category_names).split(","):
                        cat = cat.strip()
                        if not cat:
                            continue
                        category_obj = CategoryModel.objects.filter(name__iexact=cat, depth=1).first()
                        if not category_obj:
                            category_obj = CategoryModel.add_root(name=cat)
                        category_list.append(category_obj)

                    sub_category_list = []
                    for idx, subcat in enumerate(str(sub_category_names).split(",")):
                        subcat = subcat.strip()
                        if not subcat:
                            continue
                        if idx < len(category_list):
                            parent_category = category_list[idx]
                            sub_category_obj = parent_category.get_children().filter(name__iexact=subcat).first()
                            if not sub_category_obj:
                                sub_category_obj = parent_category.add_child(name=subcat)
                            sub_category_list.append(sub_category_obj)

                    if category_list:
                        product.category.set([c.id for c in category_list])
                    if sub_category_list:
                        product.sub_category.set([s.id for s in sub_category_list])

                    if created_flag:
                        created += 1
                    else:
                        updated += 1

                return Response(
                    {"status": True, "created": created, "updated": updated},
                    status=status.HTTP_200_OK
                )

            except Exception as e:
                return Response({"status": False, "error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response({'status': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    

class ProductExportView(APIView):

    def get(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return Response(
                {"status": False, "message": "Authentication credentials were not provided."},
                status=status.HTTP_401_UNAUTHORIZED
            )


        export_dir = os.path.join(settings.MEDIA_ROOT, 'export', 'products')
        os.makedirs(export_dir, exist_ok=True)


        file_name = f'products{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        file_path = os.path.join(export_dir, file_name)


        headers = [
            'Name', 'Short Name', 'Category Names', 'Sub Category Names', 'Warehouse Section',
            'Item code', 'Product Price', 'Retailer Price', 'Wholesaler Price',
             'Gst', 'Warranty', 'Unit',
            'Limited Stock', 'Out of Stock', 'Description', 'Cost', 'Weight'
        ]


        qs = ProductModel.objects.all().select_related('brand').prefetch_related('category', 'sub_category')

  
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"

 
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)


        for p in qs:
            category_names = ", ".join([c.name for c in p.category.all()])
            sub_category_names = ", ".join([s.name for s in p.sub_category.all()])
            ws.append([
                p.name,
                p.short_name,
                category_names,
                sub_category_names,
                p.warehouse_section if p.warehouse_section else "",
                p.item_code or '',
                p.product_price or '',
                p.retailer_price or '',
                p.distributer_price or '',
                p.gst or '',
                p.warranty or '',
                p.unit or '',
                p.limited_stock or '',
                p.out_of_stock or '',
                p.description or '',
                p.cost or '',
                p.weight or '',
            ])


        wb.save(file_path)


        file_uri = os.path.join(settings.MEDIA_URL, 'export', 'products', file_name)
        absolute_file_uri = request.build_absolute_uri(file_uri)

        return Response({
            'status': True,
            'file_uri': absolute_file_uri,
            'message': 'Products successfully exported'
        }, status=200)
 
class BarcodeDownloadPdfView(APIView):
    def get(self, request):
        barcode_count = request.query_params.get('barcode_count', 1)
        try:
            barcode_count = int(barcode_count)
        except ValueError:
            barcode_count = 1
        
        products = ProductModel.objects.all().order_by('-id')

        product_ids = request.query_params.get('product_ids')

        if product_ids:
            products = products.filter(id__in=ast.literal_eval(product_ids))

        product_data = []
        for product in products:
            image_url = ''    # URL to be used inside <img src="">
            # Check filesystem path if needed
            if getattr(product, 'barcode_image', None):
                # product.barcode_image.path is filesystem path
                try:
                    fs_path = product.barcode_image.path
                except Exception:
                    fs_path = None

                if fs_path and os.path.isfile(fs_path):
                    # Use the image URL (this will be like '/media/...' if using default storage)
                    # product.barcode_image.url is preferred
                    try:
                        image_url = product.barcode_image.url
                    except Exception:
                        # fallback: build from MEDIA_URL + field name
                        image_url = urljoin(settings.MEDIA_URL, product.barcode_image.name.replace(os.sep, '/'))

                    # Make image_url absolute so xhtml2pdf can fetch it if needed (useful when running behind a server)
                    # If MEDIA_URL is already absolute, this is fine; otherwise build_absolute_uri will prepend host.
                    if image_url.startswith(('http://', 'https://')):
                        absolute_image_url = image_url
                    else:
                        absolute_image_url = request.build_absolute_uri(image_url)
                else:
                    absolute_image_url = ''  # file missing
            else:
                absolute_image_url = ''

            product_info = {
                "item_code": product.item_code,
                "name": product.name,
                "barcode_image": absolute_image_url,
            }

            # 🔁 Append same barcode multiple times
            product_data.extend([product_info] * barcode_count)

        template = get_template("barcode_pdf.html")
        html = template.render({"products": product_data})

        # export_dir = os.path.join(settings.MEDIA_ROOT, "export", "product_barcodes_pdf")
        # os.makedirs(export_dir, exist_ok=True)

        # file_name = f'product_barcodes{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        # file_path = os.path.join(export_dir, file_name)

        # Create the PDF file
        # with open(file_path, "wb") as pdf_file:
        #     pisa_status = pisa.CreatePDF(html, dest=pdf_file)

        # if pisa_status.err:
        #     return Response({"status": False, "message": "Error generating PDF"}, status=500)
        
        # file_uri = os.path.join(settings.MEDIA_URL, "export", "product_barcodes_pdf", file_name)
        # absolute_file_uri = request.build_absolute_uri(file_uri)

        return Response({
            "status": True,
            "html_data": html,
            # "file_uri": absolute_file_uri,
            "message": "Product barcodes PDF successfully generated"
        }, status=200)


    
class BarcodeDownloadExcelView(APIView):
    def get(self, request):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Product Barcodes"
        row_num = 2

        # Export dir
        export_dir = os.path.join(settings.MEDIA_ROOT, "export", "products barcodes")
        os.makedirs(export_dir, exist_ok=True)

        file_name = f"product_barcodes{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = os.path.join(export_dir, file_name)

        # Header row
        sheet.append(["Item Code", "Product Name", "Barcode Image"])
        for cell in sheet[1]:
            cell.font = Font(bold=True)


        # Data rows
        for product in ProductModel.objects.all().order_by("id"):
            sheet.append([product.item_code, product.name, ""])  # placeholder for image

            sheet.row_dimensions[row_num].height = 35

            if product.barcode_image and os.path.exists(product.barcode_image.path):
                img = XLImage(product.barcode_image.path)
                img.width = 90   # adjust for better visibility
                img.height = 50
                # Insert into the "C" column (3rd column)
                sheet.add_image(img, f"C{row_num}")

            row_num += 1

        # Save file
        workbook.save(file_path)

        # Build file URI for download
        file_uri = os.path.join(settings.MEDIA_URL, "export", "products barcodes", file_name)
        absolute_file_uri = request.build_absolute_uri(file_uri)

        return Response({
            "status": True,
            "file_uri": absolute_file_uri,
            "message": "Product barcode data successfully exported"
        }, status=200)
